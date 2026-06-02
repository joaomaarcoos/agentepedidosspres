"""
Importa nomes de produtos a partir da planilha de produtos vendidos pela SPRES.

Uso:
  py execution/import_produtos_planilha.py --file "C:\\...\\LISTA.xlsx"
  py execution/import_produtos_planilha.py --file "C:\\...\\LISTA.xlsx" --apply
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(Path(__file__).resolve().parent.parent / ".env")


DEFAULT_FILIAL = "Ribeirão Preto"
BATCH_SIZE = 200


def emit(payload: dict) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False, default=str))
    sys.stdout.write("\n")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _db():
    url = os.getenv("SUPABASE_URL", "").rstrip("/")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY não configurados")
    from supabase import create_client

    return create_client(url, key)


def load_catalog(path: Path) -> dict[tuple[str, str], str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    catalog: dict[tuple[str, str], str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = _text(row[0]).upper()
        derivacao = _text(row[1]).upper()
        nome = _text(row[2]).upper()
        if cod and derivacao and nome:
            catalog[(cod, derivacao)] = nome
    return catalog


def current_products(db) -> dict[tuple[str, str], dict]:
    rows = (
        db.table("produtos")
        .select("id, filial, cod_produto, derivacao, nome, preco_base, preco_inst_299, ativo")
        .execute()
        .data
        or []
    )
    return {(_text(row.get("cod_produto")).upper(), _text(row.get("derivacao")).upper()): row for row in rows}


def table_items(db) -> list[dict]:
    return (
        db.table("tabelas_preco_itens")
        .select("id, codigo_tabela, cod_produto, variacao, nome_produto")
        .execute()
        .data
        or []
    )


def plan_changes(db, catalog: dict[tuple[str, str], str]) -> dict:
    products = current_products(db)
    items = table_items(db)

    products_to_insert = []
    products_to_update = []
    for (cod, derivacao), nome in catalog.items():
        current = products.get((cod, derivacao))
        if not current:
            products_to_insert.append(
                {
                    "filial": DEFAULT_FILIAL,
                    "cod_produto": cod,
                    "derivacao": derivacao,
                    "nome": nome,
                    "ativo": True,
                }
            )
            continue
        if _text(current.get("nome")).upper() != nome or not current.get("ativo"):
            products_to_update.append(
                {
                    "id": current["id"],
                    "cod_produto": cod,
                    "derivacao": derivacao,
                    "nome_atual": current.get("nome"),
                    "nome_novo": nome,
                    "ativo_atual": current.get("ativo"),
                }
            )

    table_items_to_update = []
    unmatched_table_items = []
    for item in items:
        key = (_text(item.get("cod_produto")).upper(), _text(item.get("variacao")).upper())
        new_name = catalog.get(key)
        if not new_name:
            if not _text(item.get("nome_produto")):
                unmatched_table_items.append(item)
            continue
        if _text(item.get("nome_produto")).upper() != new_name:
            table_items_to_update.append(
                {
                    "id": item["id"],
                    "codigo_tabela": item.get("codigo_tabela"),
                    "cod_produto": key[0],
                    "variacao": key[1],
                    "nome_atual": item.get("nome_produto"),
                    "nome_novo": new_name,
                }
            )

    return {
        "products_to_insert": products_to_insert,
        "products_to_update": products_to_update,
        "table_items_to_update": table_items_to_update,
        "unmatched_table_items": unmatched_table_items,
        "catalog_total": len(catalog),
        "products_total": len(products),
        "table_items_total": len(items),
    }


def apply_changes(db, changes: dict) -> dict:
    inserted = 0
    updated_products = 0
    updated_table_items = 0

    inserts = changes["products_to_insert"]
    for start in range(0, len(inserts), BATCH_SIZE):
        batch = inserts[start : start + BATCH_SIZE]
        if batch:
            db.table("produtos").insert(batch).execute()
            inserted += len(batch)

    for row in changes["products_to_update"]:
        db.table("produtos").update({"nome": row["nome_novo"], "ativo": True}).eq("id", row["id"]).execute()
        updated_products += 1

    for row in changes["table_items_to_update"]:
        db.table("tabelas_preco_itens").update({"nome_produto": row["nome_novo"]}).eq("id", row["id"]).execute()
        updated_table_items += 1

    return {
        "products_inserted": inserted,
        "products_updated": updated_products,
        "table_items_updated": updated_table_items,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa catálogo SPRES de uma planilha XLSX")
    parser.add_argument("--file", required=True, help="Caminho da planilha XLSX")
    parser.add_argument("--apply", action="store_true", help="Aplica as alterações no Supabase")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        emit({"ok": False, "error": f"Arquivo não encontrado: {path}"})
        return 0

    try:
        db = _db()
        catalog = load_catalog(path)
        changes = plan_changes(db, catalog)
        summary = {
            "catalog_total": changes["catalog_total"],
            "products_total_before": changes["products_total"],
            "table_items_total": changes["table_items_total"],
            "products_to_insert": len(changes["products_to_insert"]),
            "products_to_update": len(changes["products_to_update"]),
            "table_items_to_update": len(changes["table_items_to_update"]),
            "unmatched_table_items_without_name": len(changes["unmatched_table_items"]),
            "sample_products_to_insert": changes["products_to_insert"][:10],
            "sample_products_to_update": changes["products_to_update"][:10],
            "sample_table_items_to_update": changes["table_items_to_update"][:10],
            "sample_unmatched_table_items": changes["unmatched_table_items"][:10],
        }
        if args.apply:
            summary["applied"] = apply_changes(db, changes)
        else:
            summary["dry_run"] = True
        emit({"ok": True, "data": summary})
        return 0
    except Exception as exc:
        emit({"ok": False, "error": str(exc)})
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
