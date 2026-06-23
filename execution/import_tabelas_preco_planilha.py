"""
Importa itens de uma tabela de preco a partir de XLSX.

Uso:
  python execution/import_tabelas_preco_planilha.py --file "C:\\...\\TABELA 205.xlsx" --table 205
  python execution/import_tabelas_preco_planilha.py --file "C:\\...\\TABELA 205.xlsx" --table 205 --apply --prune
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

BATCH_SIZE = 200


def emit(payload: dict) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False, default=str))
    sys.stdout.write("\n")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d,.-]", "", str(value)).replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _db():
    url = os.getenv("SUPABASE_URL", "").rstrip("/")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_ANON_KEY", "")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY nao configurados")
    from supabase import create_client

    return create_client(url, key)


def _find_column(headers: list[str], candidates: tuple[str, ...]) -> int:
    normalized = [header.lower().strip() for header in headers]
    for candidate in candidates:
        candidate = candidate.lower()
        for index, header in enumerate(normalized):
            if candidate in header:
                return index
    raise ValueError(f"Coluna nao encontrada. Procurado: {', '.join(candidates)}")


def load_price_table(path: Path, table_code: str) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []

    headers = [_text(cell) for cell in rows[0]]
    cod_index = _find_column(headers, ("produto", "codigo"))
    variation_index = _find_column(headers, ("derivacao", "deriva", "variacao", "varia"))
    name_index = _find_column(headers, ("desc.prod", "descricao", "descri", "nome"))
    price_index = _find_column(headers, ("preco base", "pre", "preco"))

    synced_at = datetime.now(timezone.utc).isoformat()
    items: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for row in rows[1:]:
        cod = _text(row[cod_index] if cod_index < len(row) else "").upper()
        variation = _text(row[variation_index] if variation_index < len(row) else "").upper()
        name = _text(row[name_index] if name_index < len(row) else "").upper()
        price = _safe_float(row[price_index] if price_index < len(row) else None)
        if not cod or not variation or not name or price is None:
            continue
        key = (cod, variation)
        if key in seen:
            continue
        seen.add(key)
        items.append(
            {
                "codigo_tabela": table_code,
                "cod_produto": cod,
                "nome_produto": name,
                "variacao": variation,
                "quantidade_minima": 1,
                "preco": price,
                "desconto": 0,
                "synced_at": synced_at,
            }
        )
    return items


def current_items(db, table_code: str) -> dict[tuple[str, str], dict]:
    rows = (
        db.table("tabelas_preco_itens")
        .select("id,codigo_tabela,cod_produto,nome_produto,variacao,quantidade_minima,preco,desconto")
        .eq("codigo_tabela", table_code)
        .execute()
        .data
        or []
    )
    return {
        (_text(row.get("cod_produto")).upper(), _text(row.get("variacao")).upper()): row
        for row in rows
    }


def plan_changes(db, table_code: str, items: list[dict], prune: bool = False) -> dict:
    current = current_items(db, table_code)
    wanted = {
        (_text(item.get("cod_produto")).upper(), _text(item.get("variacao")).upper()): item
        for item in items
    }

    inserts = []
    updates = []
    for key, item in wanted.items():
        existing = current.get(key)
        if not existing:
            inserts.append(item)
            continue
        changed_fields = {}
        for field in ("nome_produto", "quantidade_minima", "preco", "desconto"):
            current_value = existing.get(field)
            wanted_value = item.get(field)
            if field == "preco":
                if round(float(current_value or 0), 4) == round(float(wanted_value or 0), 4):
                    continue
            elif _text(current_value).upper() == _text(wanted_value).upper():
                continue
            changed_fields[field] = wanted_value
        if changed_fields:
            updates.append(
                {
                    "id": existing["id"],
                    "cod_produto": key[0],
                    "variacao": key[1],
                    "antes": {
                        "nome_produto": existing.get("nome_produto"),
                        "preco": existing.get("preco"),
                    },
                    "depois": {
                        "nome_produto": item.get("nome_produto"),
                        "preco": item.get("preco"),
                    },
                    "changes": {**changed_fields, "synced_at": item.get("synced_at")},
                }
            )

    deletes = []
    if prune:
        for key, item in current.items():
            if key not in wanted:
                deletes.append(item)

    return {
        "table_code": table_code,
        "source_items": len(items),
        "current_items": len(current),
        "inserts": inserts,
        "updates": updates,
        "deletes": deletes,
    }


def apply_changes(db, table_code: str, changes: dict) -> dict:
    table_name = f"Tabela {table_code}"
    db.table("tabelas_preco").upsert(
        {
            "codigo_tabela": table_code,
            "nome_tabela": table_name,
            "synced_at": datetime.now(timezone.utc).isoformat(),
        },
        on_conflict="codigo_tabela",
    ).execute()

    inserted = 0
    inserts = changes["inserts"]
    for start in range(0, len(inserts), BATCH_SIZE):
        batch = inserts[start : start + BATCH_SIZE]
        if batch:
            db.table("tabelas_preco_itens").insert(batch).execute()
            inserted += len(batch)

    updated = 0
    for row in changes["updates"]:
        db.table("tabelas_preco_itens").update(row["changes"]).eq("id", row["id"]).execute()
        updated += 1

    deleted = 0
    for row in changes["deletes"]:
        db.table("tabelas_preco_itens").delete().eq("id", row["id"]).execute()
        deleted += 1

    return {"inserted": inserted, "updated": updated, "deleted": deleted}


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa tabela de preco XLSX para o Supabase")
    parser.add_argument("--file", required=True, help="Caminho da planilha XLSX")
    parser.add_argument("--table", required=True, help="Codigo da tabela de preco, ex: 205")
    parser.add_argument("--apply", action="store_true", help="Aplica as alteracoes")
    parser.add_argument("--prune", action="store_true", help="Remove itens que existem no banco e nao existem na planilha")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        emit({"ok": False, "error": f"Arquivo nao encontrado: {path}"})
        return 0

    try:
        table_code = args.table.strip().upper()
        db = _db()
        items = load_price_table(path, table_code)
        changes = plan_changes(db, table_code, items, prune=args.prune)
        summary = {
            "table_code": table_code,
            "source_items": changes["source_items"],
            "current_items": changes["current_items"],
            "to_insert": len(changes["inserts"]),
            "to_update": len(changes["updates"]),
            "to_delete": len(changes["deletes"]),
            "sample_inserts": changes["inserts"][:10],
            "sample_updates": changes["updates"][:10],
            "sample_deletes": changes["deletes"][:10],
        }
        if args.apply:
            summary["applied"] = apply_changes(db, table_code, changes)
        else:
            summary["dry_run"] = True
        emit({"ok": True, "data": summary})
        return 0
    except Exception as exc:
        emit({"ok": False, "error": str(exc)})
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
